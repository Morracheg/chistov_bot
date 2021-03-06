import telebot
import constants
from openpyxl import load_workbook
from time import sleep
from datetime import datetime
import logging
# TODO пока не будем отправлять
notificationsBoss = False
notificationsDev = False

rashod = dict()
rashodnik = dict()

# складывать логи в подпапку логи и TODO дробить на дни
logging.basicConfig(format=u'%(levelname)-8s [%(asctime)s] %(name)s %(message)s',
# logging.basicConfig(format=u'%(levelname)-8s [%(my_time)s] %(name)s %(message)s',
                    level=logging.INFO,
                    filename=u'логи/%s.log'%datetime.now().strftime("%d-%m-%Y"))

# telebot.logger.setLevel(logging.WARNING)
telebot.logger.setLevel(logging.INFO)

bot = telebot.TeleBot(constants.token)


def boss_message(message):
    if notificationsBoss:
        bot.send_message(constants.bossChatID, message)


def dev_message(message):
    if notificationsDev:
        bot.send_message(constants.myChatID, message)


try:
    # итак, мы имеем загруженную в память книгу
    wb = load_workbook('оборудование.xlsx')
    tools_sheet = wb['tools']
    # log_sheet = wb['log'] не используется, вместо него tools_book
    users_sheet = wb['users']
    tools_income = wb['income']  # страница журнала выручки
    tools_svod = wb['svod']  # страница журнала действий  оборудованием // 2 колонки цифр
    tools_book = wb['book']  # страница журнала действий  оборудованием
    material_book = wb['material']  # Расходник	Количество
    material_log_book = wb['material_log']  # Дата	Клинер	Материал	Движение +/-

    users = users_sheet['A']
    for user in users:
        if users_sheet.cell(row=user.row, column=3).value == 'принят':
            rashod[user.value] = False
            rashodnik[user.value] = ''

    wb2 = load_workbook('ТренингБОТ.xlsx')

except FileNotFoundError as error:
    msg = 'Не могу найти файл "%s", а без него работать не могу'%(error.filename)
    print(msg)
    # logging.error(msg, extra={'my_time': datetime.now().strftime("%H.%M")})
    logging.error(msg)
    boss_message(msg)
    raise SystemExit


def wb_save():
    global wb, wb2
    try:
        wb.save('оборудование.xlsx')
        wb2.save('ТренингБОТ.xlsx')
        return True
    except Exception as exception:
        boss_message('Не могу использовать файл эксель, исправьте и повторите')
        log_error(exception)
        return False


def log_message(message, answer):
    msg = "От %s\nТекст = %s\nОтветил = %s"%(sender_name(message),message.text,answer)
    print("\n---Вывод в консоль: Сообщение---")
    print(msg)
    # logging.info(msg, extra={'my_time': datetime.now().strftime("%H.%M")})
    logging.info(msg)
    boss_message(msg)


def log_error(err):
    print("\n---Вывод в консоль: Ошибка--- ")
    print(datetime.now().strftime("%d/%m/%y %H:%M"), '\n')
    print(err.__str__())
    # logging.error(err.__str__(), extra={'my_time': datetime.now().strftime("%H.%M")})
    logging.error(err.__str__())
    dev_message(err.__str__())


def boss_default_markup():
    default_markup = telebot.types.ReplyKeyboardMarkup(True, False)
    default_markup.row('/start', '/stop')
    default_markup.row('/Бабло', '/Люди')
    # default_markup.row('Покажи', 'Сохрани')
    default_markup.row('Отправь')
    return default_markup


def user_default_markup():
    default_markup = telebot.types.ReplyKeyboardMarkup(True, False)
    default_markup.row('Взять оборудование')
    default_markup.row('Сдать оборудование')
    default_markup.row('Взять расходники')
    default_markup.row('Обучение')
    return default_markup


def column_to_txt(letter):
    global users_sheet
    msg = ""
    for cell in users_sheet[letter]:
        # дебит и кредит не отправлять нули и только для принятых пользов "D" "E"
        # пользователей отправлять всех "C"
        if cell.row != 1:
            if letter in "DE":
                if (users_sheet.cell(row=cell.row, column=3).value == 'принят') and (int(cell.value) != 0):
                    msg += "{0} - {1}\n".format(users_sheet.cell(row=cell.row, column=1).value, cell.value)
            elif letter in "C":
                msg += "{0} - {1}\n".format(users_sheet.cell(row=cell.row, column=1).value, cell.value)
    if msg == "":
        msg = "Ничего нет!"
    return msg

# фамилия, пробел и имя отправителя сообщения
def sender_name(message):
    return message.from_user.last_name + ' ' + message.from_user.first_name


def isint(text):
    try:
        int(text)
        return True
    except ValueError:
        return False


def ispage(sheetname):
    return sheetname in wb2.sheetnames

# проверяет есть ли значение среди значений cells и возвращяет строку
def row_of_value_in_cells(value, cells):
    for cell in cells:
        if value == cell.value:
            return cell.row
    return 0


def col_row_of_value_in_cells(value, cells):
    for cell in cells:
        if value == cell.value:
            return dict(col=cell.col, row= cell.row)
    return dict(col=0, row= 0)

@bot.message_handler(commands=["start"])
def handle_text(message):
    global users_sheet
    # старт работы любого пользователя, а не когда уже начал отправлять числа
    if message.from_user.id == constants.bossChatID:
        bot.send_message(constants.bossChatID, "Работаем", reply_markup=boss_default_markup())
    elif row_of_value_in_cells(message.from_user.id, users_sheet['B']) == 0:
        # мы тут собственно регистрируем написавшего пользователя (записываем отрицательный id в переменную)
        users_sheet.append([sender_name(message), message.from_user.id, 'новенький', 0, 0])
        wb_save()
        answer = "Вас ещё нет в системе, ожидайте подтверждения"
        bot.send_message(message.chat.id, answer)
        bot.send_message(constants.bossChatID, sender_name(message) + " - новенький", reply_markup=boss_default_markup())
        log_message(message, answer)
    else:
        if users_sheet.cell(row=row_of_value_in_cells(message.from_user.id, users_sheet['B']), column=3).value == 'принят':
            bot.send_message(message.chat.id, 'Продолжим', reply_markup=user_default_markup())
        elif users_sheet.cell(row=row_of_value_in_cells(message.from_user.id, users_sheet['B']), column=3).value == 'новенький':
            bot.send_message(message.chat.id, 'Ждите')
            bot.send_message(constants.bossChatID, sender_name(message) + " - новенький и нетерпеливый",
                             reply_markup=boss_default_markup())
        logging.info('юзер шлёт start повторно, ничего не делаю')#, extra={'my_time': datetime.now().strftime("%H.%M")})


@bot.message_handler(commands=["stop"])
def handle_text(message):
    # тут юзер клавиатура убирается
    bot.send_message(message.chat.id, 'Ок, не мешаю', reply_markup=telebot.types.ReplyKeyboardRemove())


@bot.message_handler(commands=["settings"])
def handle_text(message):
    answer = "Пришла команда, но настраивать нечего!!"
    log_message(message, answer)
    bot.send_message(message.chat.id, answer)


@bot.message_handler(commands=["Бабло"])
# @bot.message_handler(func=lambda message: message.from_user.id == constants.bossChatID)
def handle_money(message):
    if message.chat.id == constants.bossChatID:
        global users_sheet
        # подтверждение кредита
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        check = False
        msg = "Подтверждение транзакций\n"
        # TODO значения в колонке E должны быть ТОЛЬКО числами
        for cell in users_sheet['E']:
            if cell.row != 1:
                if cell.value < 0:
                    username = users_sheet.cell(row=cell.row, column=1).value
                    msg += "{0}: {1}\n".format(users_sheet.cell(row=cell.row, column=1).value, cell.value)
                    user_markup.row('Получил ' + username, 'Не получил ' + username)
                    check = True
        if check:
            bot.send_message(constants.bossChatID, msg, reply_markup=user_markup)
        else:
            msg = "Долгов нет"
            bot.send_message(constants.bossChatID, msg, reply_markup=boss_default_markup())
        log_message(message, msg)


@bot.message_handler(commands=["Люди"])
# @bot.message_handler(func=lambda message: message.from_user.id == constants.bossChatID)
def handle_people(message):
    if message.chat.id == constants.bossChatID:
        # этот маркап для того, чтобы интерактивно подтверждать пользователей
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        a = 0
        for cell in users_sheet['C']:
            if cell.value == "новенький":
                username = users_sheet.cell(row=cell.row, column=1).value
                user_markup.row('Подтвердить ' + username, 'Отклонить ' + username)
                a += 1
        if a > 0:
            bot.send_message(constants.bossChatID, "Подтверждение пользователей", reply_markup=user_markup)
        else:
            bot.send_message(constants.bossChatID, "Всё в порядке", reply_markup=boss_default_markup())
        log_message(message, "Люди markup на {0} строк".format(a))


@bot.message_handler(func=lambda message: (message.from_user.id == constants.bossChatID) and constants.Boss_mode, #TODO
# @bot.message_handler(func=lambda message: message.from_user.id == constants.bossChatID,
                     content_types=["text"])
def handle_text(message):
    global users_sheet, tools_income

    if "Отмена" in message.text:
        bot.send_message(message.chat.id, "Работаем", reply_markup=boss_default_markup())
        return
    elif "Отправь" in message.text:
        if "файл" in message.text:
            if wb_save():
                bot.send_message(message.chat.id, "Файл:", reply_markup=boss_default_markup())
                doc = open('оборудование.xlsx', 'rb')
                bot.send_document(message.chat.id, doc)
            else:
                bot.send_message(message.chat.id, "Нет доступа к файлу", reply_markup=boss_default_markup())
        elif "дебит" in message.text:
            bot.send_message(message.chat.id, "Колонка дебита:", reply_markup=boss_default_markup())
            bot.send_message(constants.bossChatID, column_to_txt('D'))
        elif "кредит" in message.text:
            bot.send_message(message.chat.id, "Колонка кредита:", reply_markup=boss_default_markup())
            bot.send_message(constants.bossChatID, column_to_txt('E'))
        elif "пользовател" in message.text:
            bot.send_message(message.chat.id, "Колонка пользователей:", reply_markup=boss_default_markup())
            bot.send_message(constants.bossChatID, column_to_txt('C'))
        else:
            user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
            user_markup.row('Отправь файл', 'Отправь пользователей')
            user_markup.row('Отправь дебит', 'Отправь кредит')
            # user_markup.row('Покажи файл пользователей', 'Покажи файл с дебитом', 'Покажи файл с кредитом')
            user_markup.row('Отмена')
            bot.send_message(message.chat.id, "Что показать?", reply_markup=user_markup)

    # Босс отправляет сообщения, содержащие пользователя
    # TODO пользователь с Сохрани/Покажи/Отмена в имени сломает всё нахрен
    for cell in users_sheet['A']:
        if cell.value in message.text:
            people = True
            if "Подтвердить" in message.text:
                users_sheet.cell(row=cell.row, column=3, value='принят')
                rashod[cell.value] = False
                rashodnik[cell.value] = ''
                bot.send_message(users_sheet.cell(row=cell.row, column=2).value,
                                 "Вы подтверждены, поздравляю!\nОправляйте мне заработанные деньги "
                                 "(положительные цифры) и я буду их подсчитывать\nили сданную "
                                 "прибыль (отрицательные цифры) и я буду их списывать, если Босс "
                                 "разрешит, конечно ;)", reply_markup=user_default_markup())

            elif "Отклонить" in message.text:
                users_sheet.cell(row=cell.row, column=3, value='отклонён')
                bot.send_message(users_sheet.cell(row=cell.row, column=2).value, "Мы с Вами не дружим")

            elif "Получил" in message.text:
                from datetime import datetime
                tools_income.append([datetime.now(),
                                     users_sheet.cell(row=cell.row, column=1).value,
                                     users_sheet.cell(row=cell.row, column=5).value])

                users_sheet.cell(row=cell.row, column=4).value += users_sheet.cell(row=cell.row, column=5).value
                users_sheet.cell(row=cell.row, column=5).value = 0

                try:
                    # TODO блок try при отправке сообщений пользователю - срабытывает или нет - хз
                    bot.send_message(users_sheet.cell(row=cell.row, column=2).value, "Подтверждено")
                except telebot.apihelper.ApiException as exception:
                    log_error(exception)
                    print(exception.result)
                print('сложил', users_sheet.cell(row=cell.row, column=4).value)
                people = False

            elif "Не получил" in message.text:
                users_sheet.cell(row=cell.row, column=5).value = 0
                try:
                    # TODO блок try при отправке сообщений пользователю - срабытывает или нет - хз
                    bot.send_message(users_sheet.cell(row=cell.row, column=2).value, "Не подтверждено")
                except telebot.apihelper.ApiException as exception:
                    log_error(exception)
                    print(exception.result)
                logging.info('не сложил')#, extra={'my_time': datetime.now().strftime("%H.%M")})
                people = False

            wb_save()

            logging.info(message.text)

            if people:
                handle_people(message)
            else:
                handle_money(message)
            break

    # if "Сохрани" in message.text:
    #     if "дебит" in message.text:
    #         dict_var_save_to_file(debit, debit_save_file)
    #         bot.send_message(message.chat.id, "Хорошо", reply_markup=boss_default_markup())
    #     elif "кредит" in message.text:
    #         dict_var_save_to_file(credit, credit_save_file)
    #         bot.send_message(message.chat.id, "Хорошо", reply_markup=boss_default_markup())
    #     else:
    #         user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
    #         user_markup.row('Сохрани дебит', 'Сохрани кредит')
    #         user_markup.row('Отмена')
    #         bot.send_message(message.chat.id, "Зачем? И так всё сохраняется автоматом", reply_markup=user_markup)

    # elif "Покажи" in message.text:
    #     if "файл" in message.text:
    #         dict_var = dict()
    #         if "дебит" in message.text:
    #             dict_var_read_from_file(dict_var, debit_save_file)
    #         elif "кредит" in message.text:
    #             dict_var_read_from_file(dict_var, credit_save_file)
    #         elif "пользовател" in message.text:
    #             dict_var_read_from_file(dict_var, users_file)
    #         bot.send_message(message.chat.id, "Показываю", reply_markup=boss_default_markup())
    #         show_dict_var(dict_var)
    #     elif "дебит" in message.text:
    #         bot.send_message(message.chat.id, "Дебит:", reply_markup=boss_default_markup())
    #         show_dict_var(debit)
    #     elif "кредит" in message.text:
    #         bot.send_message(message.chat.id, "Кредит:", reply_markup=boss_default_markup())
    #         show_dict_var(credit)
    #     elif "пользовател" in message.text:
    #         bot.send_message(message.chat.id, "Пользователи:", reply_markup=boss_default_markup())
    #         show_dict_var(users)
    #     else:
    #         user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
    #         user_markup.row('Покажи пользователей', 'Покажи дебит', 'Покажи кредит')
    #         # user_markup.row('Покажи файл пользователей', 'Покажи файл с дебитом', 'Покажи файл с кредитом')
    #         user_markup.row('Отмена')
    #         bot.send_message(message.chat.id, "Что показать?", reply_markup=user_markup)

    # elif "Отмена" in message.text:
    #     bot.send_message(message.chat.id, "Работаем", reply_markup=boss_default_markup())


# TODO может убрать страницы обучения сделав их первой колонкой
# переменная для слежения за процессом обучения, содержит имя пользователя, страницу обучения и уровень глубины
level = dict()


def istested(my_dict, name):
    try:
        return my_dict[name]['level']>0
    except KeyError as e:
        log_error(e)
    return False


def unique(sequence):
    seen = set()
    return [x for x in sequence if not (x in seen or seen.add(x))]

# обработка сообщений пользователя с value == 'принят'
# TODO обработка сообщений таким образом пропускает всех, кто начал не со "/start"
@bot.message_handler(func=lambda message: (message.from_user.id != constants.bossChatID or constants.Worker_mode) and #TODO
# @bot.message_handler(func=lambda message: message.from_user.id != constants.bossChatID and
                                          users_sheet.cell(row=row_of_value_in_cells(message.from_user.id, users_sheet['B']), column=3).value == 'принят',
                     content_types=["text"])
def handle_text(message):
    global users_sheet, tools_sheet, tools_book, tools_income # log_sheet
    global rashod, material_book, rashodnik, material_log_book, level
    if 'Взять оборудование' == message.text:
        # отправить список незанятого оборудования
        tools_tools = tools_sheet['A']  # колонка наименования оборудования

        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        check = False  # есть ли вообще доступное оборудование
        for tool in tools_tools:
            if tool.value is not None \
                    and tools_sheet.cell(row=tool.row, column=2).value == 'Свободен' \
                    and tool.row != 1:
                user_markup.row('Взял {0}'.format(tool.value))
                check = True
        user_markup.row('Отмена')

        if check:  # оборудование есть
            bot.send_message(message.chat.id, "Вот список доступного оборудования",
                             reply_markup=user_markup)
        else:  # оборудования нет
            bot.send_message(message.chat.id, "К сожалению сейчас нет доступного оборудования",
                             reply_markup=user_default_markup())
            bot.send_message(constants.bossChatID,
                             '{0} хотел взять оборудование, но его разобрали!'.format(sender_name(message)))

    elif 'Взять расходники' == message.text:
        rashod[sender_name(message)] = True
        # №1 Клинеру выпадает список доступных расходников
        # После этого сальдо по нему уменьшается.
        materials = material_book['A']  # колонка наименования расходника
        msg = ''
        check = False
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        for material in materials:
            if material.value is not None \
                    and material.row != 1:
                    # and int(material_book.cell(row=material.row, column=2).value) > 0 \
                msg += '{0}: {1} единиц\n'.format(material.value, material_book.cell(row=material.row, column=2).value)
                user_markup.row('Взял {0}'.format(material.value))
                check = True
        user_markup.row('Отмена')

        if check:  # расходники есть
            bot.send_message(message.chat.id, msg, reply_markup=user_markup)
        else:  # расходников нет
            bot.send_message(message.chat.id, "Вы не можете ничего взять",
                             reply_markup=user_default_markup())
            bot.send_message(constants.bossChatID,
                             '{0} хотел взять расходник, но его разобрали!'.format(sender_name(message)))

    elif 'Сдать оборудование' == message.text:
        # отправить список взятого оборудования
        tools_tools = tools_sheet['A']  # колонка оборудования

        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        check = False  # есть ли взятое человеком оборудование
        for tool in tools_tools:
            if tool.value is not None \
                    and tools_sheet.cell(row=tool.row, column=2).value == 'Взят' \
                    and tools_sheet.cell(row=tool.row, column=3).value == sender_name(message):
                user_markup.row('Сдал {0}'.format(tool.value))
                check = True
        user_markup.row('Отмена')

        if check:  # оборудование есть
            bot.send_message(message.from_user.id, "Вот список оборудования у Вас на руках",
                             reply_markup=user_markup)
        else:  # оборудования нет
            bot.send_message(message.from_user.id, "Так Вы ничего и не брали",
                             reply_markup=user_default_markup())

    elif 'Взял' in message.text or 'Сдал' in message.text:
        # пришло (а пришло ли?) название оборудования и слово Взял или Сдал
        # TODO ещё пополнить счетчик уборок? +1 когда вернули (но не всегда же +1)
        # TODO ещё счетчик времени уборок? суммировать время нахождения оборудования у клинера?
        # TODO опять же не всегда он будет его сразу после уборки сдавать
        # TODO ещё возможность написать примечание о конкретном оборудовании?

        if rashod[sender_name(message)]:
            # №2 Он выбирает другая переменная запоминает его выбор и ему предлагается ввести сколько он берет этого расходника.
            rashodnik[sender_name(message)] = message.text
            user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
            user_markup.row('Отмена')
            bot.send_message(message.chat.id, 'Сколько?', reply_markup=user_markup)
        else:
            from datetime import datetime

            tools_tools = tools_sheet['A']  # колонка оборудования
            for tool in tools_tools:
                if tool.value in message.text:  # добавить в статус(колонка 2) Взят или свободен
                    if 'Взял' in message.text:
                        tools_sheet.cell(row=tool.row, column=2).value = 'Взят'
                        # если Взят то добавить кем взят в 3ю колонку
                        tools_sheet.cell(row=tool.row, column=3).value = sender_name(message)
                        tools_book.append([datetime.now(), sender_name(message), tool.value, 'Взяла'])
                    elif 'Сдал' in message.text:
                        tools_sheet.cell(row=tool.row, column=2).value = 'Свободен'
                        tools_sheet.cell(row=tool.row, column=3).value = 'На базе'
                        tools_book.append([datetime.now(), sender_name(message), tool.value, 'Сдала'])
                    # записать в колонку 4 дату этого (любого?) события
                    tools_sheet.cell(row=tool.row, column=4).value = datetime.now()

            wb_save()

            bot.send_message(message.from_user.id, "Принято", reply_markup=user_default_markup())
            bot.send_message(constants.bossChatID, '{0}: "{1}"'.format(sender_name(message), message.text))

    elif isint(message.text):
        if rashod[sender_name(message)]:
            if int(message.text) > 0:
                check = False
                materials = material_book['A']
                for material in materials:
                    if material.value in rashodnik[sender_name(message)]:
                        if sender_name(message) in 'Gladneva Inna':
                            material_book.cell(row=material.row, column=2).value += int(message.text)
                        else:
                            material_book.cell(row=material.row, column=2).value -= int(message.text)
                        check = True
                        if int(material_book.cell(row=material.row, column=2).value) <= int(
                                material_book.cell(row=material.row, column=3).value):
                            bot.send_message(constants.bossChatID,
                                             'Босс пора пополнить запас {0}'.format(material.value))

                if check:
                    msg = 'Принято: "{0} {1}"'.format(rashodnik[sender_name(message)], message.text)
                    logging.info('{0}: {1}'.format(sender_name(message), msg))
                    bot.send_message(message.chat.id, msg, reply_markup=user_default_markup())

                    #rashodnik[sender_name(message)] = ''
                    #rashod[sender_name(message)] = False

                    bot.send_message(constants.bossChatID,
                                     '{0} отправил: "{1} {2}" (расходник)'.format(sender_name(message),
                                                                                  rashodnik[sender_name(message)],
                                                                                  message.text))

                    # запишем в лог
                    from datetime import datetime
                    # здесь реверсивно
                    up_or_down = '+' if (sender_name(message) in 'Gladneva Inna') else '-'
                    up_or_down += message.text
                    material_log_book.append(
                        [datetime.now(), sender_name(message), rashodnik[sender_name(message)], up_or_down])
                    rashodnik[sender_name(message)] = ''
                    rashod[sender_name(message)] = False

                    wb_save()

                else:
                    logging.info('Наёбывают')
                    rashodnik[sender_name(message)] = ''
                    rashod[sender_name(message)] = False
                    bot.send_message(message.chat.id, 'Какая-то ошибка, давайте заново', reply_markup=user_default_markup())
            else:
                user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
                user_markup.row('Отмена')
                bot.send_message(message.chat.id, 'Отправьте положительное число\nТак сколько?', reply_markup=user_markup)

        else:
            bot.send_message(constants.bossChatID, '{0} отправил: "{1}"(выручка)'.format(sender_name(message), message.text))
            answer = ''
            if int(message.text) < 0:
                users_sheet.cell(row=row_of_value_in_cells(message.from_user.id, users_sheet['B']), column=5).value += int(message.text)
                # answer = "Принял, требуется подтверждение"
                answer = 'Принято: "В кассу {0} рублей". Требуется подтверждение босса.'.format(abs(int(message.text)))
                logging.info('Добавил в кредит: ', '{0}: "{1}"'.format(sender_name(message), message.text))
            else:
                from datetime import datetime
                tools_income.append([datetime.now(), sender_name(message), message.text])

                users_sheet.cell(row=row_of_value_in_cells(message.from_user.id, users_sheet['B']), column=4).value += int(message.text)
                answer = 'Принято: "На руки {0} рублей"'.format(int(message.text))
                logging.info('Добавил в дебит: ', '{0}: "{1}"'.format(sender_name(message), message.text))
            wb_save()
            log_message(message, answer)
            bot.send_message(message.chat.id, answer, reply_markup=user_default_markup())

    elif 'Отмена' == message.text:
        rashodnik[sender_name(message)] = ''
        rashod[sender_name(message)] = False
        answer = "Работаем!"
        log_message(message, answer)
        bot.send_message(message.chat.id, answer, reply_markup=user_default_markup())

    elif 'Обучение' == message.text:
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        for i in range(len(wb2.sheetnames) - 2):
            user_markup.row(wb2.sheetnames[i])
        user_markup.row('Отмена')
        bot.send_message(message.from_user.id, "Выбирайте интересующие вас разделы, и изучайте информацию.",
                         reply_markup=user_markup)
        level.update({sender_name(message): {'sheet': message.text, 'level': 0}})

    elif ispage(message.text):
        level.update({sender_name(message): {'sheet': message.text, 'level': 1}})
        # TODO начать таймер

        A_column = wb2[level[sender_name(message)]['sheet']]['A']

        my_list = list()
        for cell in A_column:
            assert cell.value != None, "неправильная ячейка: {0} {1} = None".format(cell.row, cell.column)
            if cell.value not in my_list:
                my_list.append(cell.value)

        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        # for item in unique(my_list):
        for item in my_list:
            assert isinstance(item, str), 'oops not a string - {0}'.format(type(item))
            user_markup.row(item)
        user_markup.row('Отмена')
        # if len(testsheet.columns) != 2:
        bot.send_message(message.from_user.id, "Обучение. Выбирайте что хотите узнать.",
                         reply_markup=user_markup)
        #если второй столбец не последни заполненный из выбранного раздела,
        #  то бот вываливает кнопки с уникальными разделами второго столбца и так далее

    elif istested(level, sender_name(message)):
        # {sender_name(message): {'sheet': message.text, 'level': 1}}
        my_sheet = level[sender_name(message)]['sheet']
        my_col = level[sender_name(message)]['level']

        N_sheet = wb2[my_sheet]

        from string import ascii_uppercase
        assert my_col <= len(ascii_uppercase), "oh shi too deep"
        N_column = N_sheet[ascii_uppercase[my_col-1]]

        if row_of_value_in_cells(message.text, N_column):
            my_row = row_of_value_in_cells(message.text, N_column)
        else:
            bot.send_message(message.from_user.id,
                             "Ошибка ввода. Давайте всё по новой",
                             reply_markup=user_default_markup())
            del level[sender_name(message)]
            return

        if N_sheet.cell(row=my_row, column=my_col+2).value == None:
            for cell in N_column:
                if cell.value == message.text:
                    # assert (N_sheet.cell(row=cell.row, column=my_col-1).value == \
                    #     N_sheet.cell(row=my_row, column=my_col-1).value), 'неправильная выдача'
                    # bot.send_message(message.from_user.id,
                    #                  N_sheet.cell(row=cell.row, column=my_col+1).value,
                    #                  reply_markup=user_default_markup())
                    if my_col == 1:
                        bot.send_message(message.from_user.id,
                                         N_sheet.cell(row=cell.row, column=my_col + 1).value,
                                         reply_markup=user_default_markup())
                    elif (N_sheet.cell(row=cell.row, column=my_col-1).value == N_sheet.cell(row=my_row, column=my_col-1).value):
                        bot.send_message(message.from_user.id,
                                         N_sheet.cell(row=cell.row, column=my_col + 1).value,
                                         reply_markup=user_default_markup())
                    else:
                        assert False, N_sheet.cell(row=cell.row, column=my_col-1).value
            del level[sender_name(message)]
        else:
            my_list = list()
            for cell in N_column:
                if cell.value == message.text and N_sheet.cell(row=cell.row, column=my_col+1).value not in my_list:
                    my_list.append(N_sheet.cell(row=cell.row, column=my_col+1).value)

            user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
            for item in my_list:
                user_markup.row(item)
            user_markup.row('Отмена')

            bot.send_message(message.from_user.id, "Обучение. Выбирайте что хотите узнать.",
                             reply_markup=user_markup)

            # my_col = my_col+1
            level[sender_name(message)]['level'] = level[sender_name(message)]['level']+1


        #проверка наличия в листе
        # colrow_dict = col_row_of_value_in_cells(message.text, wb2[level[sender_name(message)]['sheet']].get_cell_collection)
        # if colrow_dict['col'] != 0 and colrow_dict['row'] !=0:
        #     # N_column = wb2[level[sender_name(message)]['sheet']].iter_cols()
        #     N_sheet = wb2[level[sender_name(message)]['sheet']]
        #     N_column = N_sheet.columns[colrow_dict['col']]
        #
        #     user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        #     # мы нашли на странице значение и должны показать или кнопки или обучающую информацию
        #     if N_sheet.cell(row=colrow_dict['row'], col=colrow_dict['col']+2).value == None:
        #         bot.send_message(message.from_user.id,
        #                          N_sheet.cell(row=colrow_dict['row'], col=colrow_dict['col']+1).value,
        #                          reply_markup=user_default_markup())
        #         del level[sender_name(message)]
        #         for cell in N_column:
        #             pass
        #         # tools_sheet.cell(row=tool.row, column=2).value
        #
        #     user_markup.row('Отмена')
        #
        #     bot.send_message(message.from_user.id, "Обучение. Выбирайте что хотите узнать.",
        #                      reply_markup=user_markup)

    else:
        answer = "Что-что?"
        log_message(message, answer)
        bot.send_message(message.chat.id, answer, reply_markup=user_default_markup())


# value='отклонён'
@bot.message_handler(func=lambda message: message.from_user.id != constants.bossChatID and
                                          users_sheet.cell(row=row_of_value_in_cells(message.from_user.id, users_sheet['B']), column=3).value == 'отклонён',
                     content_types=["text"])
def handle_text(message):
    log_message(message, "Заблокированный юзер балуется: " + sender_name(message))


# value='новенький'
@bot.message_handler(func=lambda message: message.from_user.id != constants.bossChatID and
                                          users_sheet.cell(row=row_of_value_in_cells(message.from_user.id, users_sheet['B']), column=3).value == 'новенький',
                     content_types=["text"])
def handle_text(message):
    log_message(message, "Неподтвержденный юзер шлёт чото")
    bot.send_message(message.chat.id, "Вы не подтверждены, ждите, Босс обдумывает вашу кандидатуру")
    bot.send_message(constants.bossChatID,
                     'Неподтвержденный {0} отправил: "{1}"'.format(sender_name(message), message.text))


@bot.message_handler(content_types=["document"])
def handle_text(message):
    bot.send_message(message.chat.id, "Пришел документ, неожиданно /start")


@bot.message_handler(content_types=["audio"])
def handle_text(message):
    bot.send_message(message.chat.id, "Пришла аудиозапись, неожиданно /start")


@bot.message_handler(content_types=["photo"])
def handle_text(message):
    bot.send_message(message.chat.id, "Пришло изображение, неожиданно /start")


@bot.message_handler(content_types=["sticker"])
def handle_text(message):
    bot.send_message(message.chat.id, "Пришел стикер, неожиданно /start")


if __name__ == "__main__":
    while True:
        try:
            logging.info('Запустил')#, extra={'my_time': datetime.now().strftime("%H.%M")})
            bot.polling(none_stop=True, timeout=60)
            # bot.polling(none_stop=True, interval=0)
        except Exception as e:
            wb_save()
            bot.stop_polling()
            log_error(e)
            sleep(15)
